#__author__ = 'chubby_superman'
#_*_coding=utf-8 _*

import xlwt
import openpyxl
import hack_test
data_dict= [
{
 "text":"sf",
 "id": 1234,
 "detail":hack_test.sku()[1][0]
}
,
{
 "text":"MXCHIP won a prize",
 "id": 2345,
 "detail":["asdfadfa",[],{}]
},
]
datas=data_dict
def write_excel(datas,name):
    workbook=openpyxl.Workbook()
    sheet = workbook.active
    c = 0
    for i in range(len(datas)):
        d = datas[i]
        title = list(d.keys())
        for i in range(len(title)):
            sheet.cell(row = 1,column =i+1,value =title[i])
        l = [d[k] for k in title]
        c +=1
        for k1 in range(len(l)):
            sheet.cell(row = c+1,column =k1+1,value ="{}".format(l[k1]))
    workbook.save(name)
write_excel(datas,"jd.xlsx")